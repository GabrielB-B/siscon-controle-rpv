from __future__ import annotations

import argparse
import copy
from collections import Counter
from datetime import date, datetime
from decimal import Decimal
import os
from pathlib import Path

from openpyxl import load_workbook
from unidecode import unidecode


APPROVED_PENDINGS = {
    "TIPO ORIGINAL RPV-DANOS MORAIS MAPEADO PARA INDENIZACAO",
}

# Linhas da aba de pendencias que foram liberadas manualmente para cadastro
# mesmo nao estando no lote seguro original da auditoria.
APPROVED_PENDING_ROWS = {
    83: "C.I. numerica interpretada como data; cadastro liberado com ajuste manual",
}

# Casos em que a C.I. foi lida como data na planilha original.
PROCESSO_EDOC_OVERRIDES = {
    83: "01/2026",
    91: "06/2026",
}

# No caso da linha 83 a planilha original nao traz a DATA_CI em coluna propria.
# Como a liberacao manual foi dada com base na referencia mensal da C.I. 01/2026,
# usamos o primeiro dia do mes como data operacional de fallback para permitir o cadastro.
DATA_CI_OVERRIDES = {
    83: date(2026, 1, 1),
}


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Importa o subconjunto seguro de RPVs normais a partir da auditoria final."
    )
    parser.add_argument(
        "--source-xlsx",
        default="entrada/rpvs_normais/Pasta1.xlsx",
        help="Planilha Excel original usada como fonte de dados completos.",
    )
    parser.add_argument(
        "--audit-xlsx",
        default=(
            "saida/rpvs_normais/"
            "auditoria_final_cadastro_manual_2026-03-31_094524/"
            "RPVS_NORMAIS_AUDITORIA_PASTA1.xlsx"
        ),
        help="Planilha de auditoria final.",
    )
    parser.add_argument(
        "--db-path",
        default="instance/controle_rpv.db",
        help="Banco SQLite alvo.",
    )
    parser.add_argument(
        "--output-dir",
        default="saida/rpvs_normais",
        help="Pasta base para os relatórios desta importação assistida.",
    )
    parser.add_argument(
        "--backup-dir",
        default="",
        help="Pasta de backup usada antes da execução, para registrar no relatório.",
    )
    parser.add_argument(
        "--execute",
        action="store_true",
        help="Executa a importação no banco. Sem esta flag, roda apenas a prévia.",
    )
    return parser.parse_args()


def _nk(value: object) -> str:
    return unidecode(str(value or "").strip()).upper()


def _texto(value: object) -> str:
    return str(value or "").strip()


def _decimal_chave(value: object) -> str:
    return f"{Decimal(str(value or 0)):.2f}"


def _ler_linhas_aprovadas(audit_path: Path) -> tuple[dict[int, str], list[dict[str, object]]]:
    workbook = load_workbook(audit_path, read_only=True, data_only=True)
    try:
        selected: dict[int, str] = {}
        pendentes_restantes: list[dict[str, object]] = []

        dup_sheet = workbook["DUPLICADOS_16"]
        dup_rows = list(dup_sheet.iter_rows(values_only=True))
        dup_headers = {name: index for index, name in enumerate(dup_rows[0])}
        for row in dup_rows[1:]:
            linha_original = int(row[dup_headers["LINHA_ORIGINAL"]])
            selected[linha_original] = "Duplicidade validada na auditoria final"

        faltam_sheet = workbook["FALTAM_MANUAL_56"]
        faltam_rows = list(faltam_sheet.iter_rows(values_only=True))
        faltam_headers = {name: index for index, name in enumerate(faltam_rows[0])}

        for row in faltam_rows[1:]:
            linha_original = int(row[faltam_headers["LINHA_ORIGINAL"]])
            if linha_original in selected:
                continue

            if linha_original in APPROVED_PENDING_ROWS:
                selected[linha_original] = APPROVED_PENDING_ROWS[linha_original]
                continue

            pendencias = _texto(row[faltam_headers["PENDENCIAS"]])
            pendencias_norm = _nk(pendencias)

            if pendencias_norm in APPROVED_PENDINGS:
                selected[linha_original] = "Mapeamento de tipo aprovado para cadastro"
                continue

            pendentes_restantes.append(
                {
                    "linha_original": linha_original,
                    "processo_edoc": _texto(row[faltam_headers["PROCESSO_EDOC"]]),
                    "numero_processo": _texto(row[faltam_headers["NUMERO_PROCESSO"]]),
                    "nome_beneficiario": _texto(row[faltam_headers["NOME_BENEFICIARIO"]]),
                    "motivo_principal": _texto(row[faltam_headers["MOTIVO_PRINCIPAL"]]),
                    "pendencias": pendencias,
                    "acao_recomendada": _texto(row[faltam_headers["ACAO_RECOMENDADA"]]),
                }
            )

        return selected, pendentes_restantes
    finally:
        workbook.close()


def _preparar_linhas_importacao(
    *,
    source_path: Path,
    selected_rows: dict[int, str],
) -> tuple[list[object], list[dict[str, object]]]:
    from app.services.rpv_import_service import carregar_planilha_rpvs_normais

    linhas_fonte = {
        linha.linha_original: linha
        for linha in carregar_planilha_rpvs_normais(source_path)
    }

    selecionadas: list[object] = []
    relatorio: list[dict[str, object]] = []

    for linha_original, motivo in sorted(selected_rows.items()):
        origem = linhas_fonte.get(linha_original)
        if not origem:
            raise ValueError(f"Linha {linha_original} não encontrada na planilha de origem.")

        linha = copy.deepcopy(origem)

        if linha_original in PROCESSO_EDOC_OVERRIDES:
            linha.processo_edoc = PROCESSO_EDOC_OVERRIDES[linha_original]
        if linha_original in DATA_CI_OVERRIDES:
            linha.data_ci = DATA_CI_OVERRIDES[linha_original]

        linha.issues = []
        linha.duplicado_banco = False
        linha.duplicado_planilha = False
        linha.duplicado_detalhe = None
        linha.registro_existente_id = None
        linha.conciliavel_banco = False
        linha.conciliado = False
        linha.conciliacao_detalhe = None
        linha.importado = False
        linha.registro_id = None
        linha.erro_importacao = None

        campos_obrigatorios = {
            "exercicio": linha.exercicio,
            "processo_edoc": linha.processo_edoc,
            "nome_beneficiario": linha.nome_beneficiario,
            "documento_ajustado": linha.documento_ajustado,
            "tipo_documento": linha.tipo_documento,
            "numero_processo": linha.numero_processo,
            "numero_processo_normalizado": linha.numero_processo_normalizado,
            "data_ci": linha.data_ci,
            "valor_bruto": linha.valor_bruto,
            "tipo_rpv_destino": linha.tipo_rpv_destino,
            "elaborador_destino": linha.elaborador_destino,
            "situacao_empenho_destino": linha.situacao_empenho_destino,
            "situacao_imposto_destino": linha.situacao_imposto_destino,
        }
        faltantes = [nome for nome, valor in campos_obrigatorios.items() if not valor]
        if faltantes:
            raise ValueError(
                f"Linha {linha_original} ainda não está segura para importação. "
                f"Campos faltantes: {', '.join(faltantes)}"
            )

        selecionadas.append(linha)
        relatorio.append(
            {
                "linha_original": linha.linha_original,
                "motivo_aprovacao": motivo,
                "processo_edoc": linha.processo_edoc,
                "numero_processo": linha.numero_processo,
                "nome_beneficiario": linha.nome_beneficiario,
                "documento": linha.documento_ajustado,
                "tipo_rpv": linha.tipo_rpv_destino,
                "valor_bruto": linha.valor_bruto,
            }
        )

    return selecionadas, relatorio


def _validar_contra_banco(linhas: list[object]) -> None:
    from app.services.rpv_import_service import aplicar_bloqueios_banco, coletar_processos_existentes

    aplicar_bloqueios_banco(
        linhas,
        chaves_existentes=coletar_processos_existentes(),
    )

    conflitos = [linha for linha in linhas if linha.duplicado_banco]
    if conflitos:
        detalhes = [
            f"linha {linha.linha_original}: {linha.duplicado_detalhe}"
            for linha in conflitos
        ]
        raise ValueError(
            "Foram encontrados conflitos novos contra o banco atual: "
            + " | ".join(detalhes)
        )


def _separar_ja_existentes(linhas: list[object]) -> tuple[list[object], list[dict[str, object]]]:
    from app.models import Processo, RegistroRPV
    from app.utils.normalizers import (
        normalizar_documento,
        normalizar_nome,
        normalizar_numero_processo,
    )

    registros = RegistroRPV.query.join(Processo).all()
    restantes: list[object] = []
    ja_existentes: list[dict[str, object]] = []

    for linha in linhas:
        chave_processo = normalizar_numero_processo(linha.numero_processo)
        chave_nome = normalizar_nome(linha.nome_beneficiario)
        chave_doc = normalizar_documento(linha.documento_ajustado)
        chave_valor = _decimal_chave(linha.valor_bruto)
        chave_tipo = _nk(linha.tipo_rpv_destino)

        registro_encontrado = None
        for registro in registros:
            processo = registro.processo
            if not processo or not registro.ativo:
                continue
            if normalizar_numero_processo(processo.numero_processo) != chave_processo:
                continue
            if normalizar_nome(registro.nome_beneficiario) != chave_nome:
                continue
            if normalizar_documento(registro.documento_original) != chave_doc:
                continue
            if _decimal_chave(registro.valor_bruto) != chave_valor:
                continue
            if _nk(getattr(getattr(registro, "tipo_rpv", None), "nome", None)) != chave_tipo:
                continue

            registro_encontrado = registro
            break

        if registro_encontrado:
            ja_existentes.append(
                {
                    "linha_original": linha.linha_original,
                    "registro_id": registro_encontrado.id,
                    "processo_edoc": linha.processo_edoc,
                    "numero_processo": linha.numero_processo,
                    "nome_beneficiario": linha.nome_beneficiario,
                }
            )
            continue

        restantes.append(linha)

    return restantes, ja_existentes


def _escrever_relatorio(
    *,
    output_dir: Path,
    source_path: Path,
    audit_path: Path,
    db_path: Path,
    backup_dir: Path | None,
    execute: bool,
    linhas_planejadas: list[dict[str, object]],
    linhas_ja_existentes: list[dict[str, object]],
    linhas_importadas: list[dict[str, object]],
    pendentes_restantes: list[dict[str, object]],
) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)

    report_path = output_dir / "relatorio_importacao_assistida.md"
    totais = Counter()
    totais["planejadas"] = len(linhas_planejadas)
    totais["ja_existentes"] = len(linhas_ja_existentes)
    totais["importadas"] = len(linhas_importadas)
    totais["pendentes_restantes"] = len(pendentes_restantes)

    linhas_md = [
        "# Importação assistida dos RPVs restantes",
        "",
        f"- Origem: `{source_path}`",
        f"- Auditoria usada: `{audit_path}`",
        f"- Banco alvo: `{db_path}`",
        f"- Backup prévio: `{backup_dir}`" if backup_dir else "- Backup prévio: não informado",
        f"- Modo: {'execução' if execute else 'prévia'}",
        f"- Linhas planejadas para este lote seguro: {totais['planejadas']}",
        f"- Já encontradas no banco e puladas nesta rodada: {totais['ja_existentes']}",
        f"- Linhas efetivamente importadas: {totais['importadas']}",
        f"- Linhas ainda pendentes para ação manual: {totais['pendentes_restantes']}",
        "",
        "## Linhas deste lote seguro",
    ]

    for item in linhas_planejadas:
        linhas_md.append(
            "- "
            f"linha {item['linha_original']} | "
            f"CI {item['processo_edoc']} | "
            f"processo {item['numero_processo']} | "
            f"{item['nome_beneficiario']} | "
            f"{item['motivo_aprovacao']}"
        )

    linhas_md.extend(
        [
            "",
            "## Já estavam no sistema",
        ]
    )

    if linhas_ja_existentes:
        for item in linhas_ja_existentes:
            linhas_md.append(
                "- "
                f"linha {item['linha_original']} já corresponde ao registro {item['registro_id']} | "
                f"CI {item['processo_edoc']} | "
                f"processo {item['numero_processo']} | "
                f"{item['nome_beneficiario']}"
            )
    else:
        linhas_md.append("- Nenhuma das linhas planejadas já estava no banco.")

    linhas_md.extend(
        [
            "",
            "## Importadas nesta execução",
        ]
    )

    if linhas_importadas:
        for item in linhas_importadas:
            linhas_md.append(
                "- "
                f"linha {item['linha_original']} -> registro {item['registro_id']} | "
                f"CI {item['processo_edoc']} | "
                f"processo {item['numero_processo']} | "
                f"{item['nome_beneficiario']}"
            )
    else:
        linhas_md.append("- Nenhuma linha foi importada porque a execução foi apenas de prévia.")

    linhas_md.extend(
        [
            "",
            "## Pendentes que continuam fora",
        ]
    )

    for item in pendentes_restantes:
        linhas_md.append(
            "- "
            f"linha {item['linha_original']} | "
            f"CI {item['processo_edoc'] or '-'} | "
            f"processo {item['numero_processo'] or '-'} | "
            f"{item['nome_beneficiario'] or '-'} | "
            f"{item['pendencias'] or item['motivo_principal']}"
        )

    report_path.write_text("\n".join(linhas_md), encoding="utf-8")
    return report_path


def main() -> int:
    args = _parse_args()

    source_path = Path(args.source_xlsx).resolve()
    audit_path = Path(args.audit_xlsx).resolve()
    db_path = Path(args.db_path).resolve()
    output_root = Path(args.output_dir).resolve()
    backup_dir = Path(args.backup_dir).resolve() if args.backup_dir else None

    if not source_path.exists():
        print(f"Planilha de origem não encontrada: {source_path}")
        return 1

    if not audit_path.exists():
        print(f"Planilha de auditoria não encontrada: {audit_path}")
        return 1

    if not db_path.exists():
        print(f"Banco alvo não encontrado: {db_path}")
        return 1

    os.environ["DATABASE_URL"] = f"sqlite:///{db_path.as_posix()}"

    from app import create_app
    from app.services.rpv_import_service import importar_linhas

    stamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    output_dir = output_root / f"importacao_assistida_auditoria_{stamp}"

    selected_rows, pendentes_restantes = _ler_linhas_aprovadas(audit_path)

    app = create_app()
    with app.app_context():
        linhas, linhas_planejadas = _preparar_linhas_importacao(
            source_path=source_path,
            selected_rows=selected_rows,
        )
        linhas, ja_existentes = _separar_ja_existentes(linhas)
        _validar_contra_banco(linhas)

        importadas: list[dict[str, object]] = []
        if args.execute:
            resultado = importar_linhas(linhas)
            if resultado.get("erros", 0):
                raise RuntimeError(
                    f"A importação assistida terminou com erros: {resultado.get('erros', 0)}"
                )

            for linha in linhas:
                if linha.importado and linha.registro_id:
                    importadas.append(
                        {
                            "linha_original": linha.linha_original,
                            "registro_id": linha.registro_id,
                            "processo_edoc": linha.processo_edoc,
                            "numero_processo": linha.numero_processo,
                            "nome_beneficiario": linha.nome_beneficiario,
                        }
                    )

        report_path = _escrever_relatorio(
            output_dir=output_dir,
            source_path=source_path,
            audit_path=audit_path,
            db_path=db_path,
            backup_dir=backup_dir,
            execute=args.execute,
            linhas_planejadas=linhas_planejadas,
            linhas_ja_existentes=ja_existentes,
            linhas_importadas=importadas,
            pendentes_restantes=pendentes_restantes,
        )

    print(f"Linhas seguras selecionadas: {len(linhas_planejadas)}")
    print(f"Relatório: {report_path}")
    if args.execute:
        print(f"Importadas nesta execução: {len(importadas)}")
    else:
        print("Prévia concluída sem gravação no banco.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
