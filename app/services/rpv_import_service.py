from __future__ import annotations

from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation
from pathlib import Path
import re

from openpyxl import Workbook, load_workbook
from unidecode import unidecode

from app.extensions import db
from app.models import (
    DativoItem,
    Processo,
    RegistroRPV,
    SituacaoEmpenho,
    SituacaoImposto,
    TipoRPV,
    User,
)
from app.services.audit_service import registrar_evento, snapshot_entidade
from app.utils.documentos import validar_documento_brasileiro
from app.utils.normalizers import (
    normalizar_documento,
    normalizar_nome,
    normalizar_numero_processo,
)


SKIP_DESCRICOES = {"DATIVO", "RPV-DATIVO", "RPV DATIVO"}
DOCUMENTO_PLACEHOLDERS = {"--", "-", "XXX", "XXXXXXXXXXX"}
NOME_PLACEHOLDERS = {"--", "-", ""}
STATUS_EMPENHO_PAGOS = {"PAGO", "CONCLUIDA"}


def _nk(value) -> str:
    return unidecode(str(value or "").strip()).upper()


def _nk_compacto(value) -> str:
    return re.sub(r"\s+", " ", _nk(value))


def _texto_planilha(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    texto = str(value).strip()
    if texto.endswith(".0") and texto[:-2].isdigit():
        return texto[:-2]
    return texto


def _texto_limpo(value) -> str | None:
    texto = _texto_planilha(value).strip()
    return texto or None


def _decimal_planilha(value):
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))

    texto = str(value).strip()
    if not texto:
        return None
    if texto.casefold() in {"s/if", "sif"}:
        return texto

    texto = texto.replace("R$", "").replace(" ", "")
    if "," in texto and "." in texto:
        texto = texto.replace(".", "").replace(",", ".")
    else:
        texto = texto.replace(",", ".")

    try:
        return Decimal(texto)
    except InvalidOperation:
        return texto


def _competencia_planilha(value) -> str | None:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m")
    if isinstance(value, date):
        return value.strftime("%Y-%m")

    texto = str(value or "").strip()
    if not texto:
        return None

    match = re.match(r"^(\d{4})[-/](\d{2})$", texto)
    if match:
        return f"{match.group(1)}-{match.group(2)}"

    match = re.match(r"^(\d{2})[-/](\d{4})$", texto)
    if match:
        return f"{match.group(2)}-{match.group(1)}"

    return None


def _data_planilha(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    texto = str(value or "").strip()
    if not texto:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(texto, fmt).date()
        except ValueError:
            continue
    return None


def _data_base_competencia(competencia: str | None):
    if not competencia or len(competencia) != 7 or "-" not in competencia:
        return None
    ano, mes = competencia.split("-")
    return date(int(ano), int(mes), 1)


def _normalizar_documento_importacao(raw) -> tuple[str | None, str | None, str, str | None]:
    texto_original = _texto_limpo(raw)
    if not texto_original or texto_original in DOCUMENTO_PLACEHOLDERS:
        return None, None, "PENDENTE_SEM_DOCUMENTO", None

    digitos = re.sub(r"\D", "", texto_original)
    if not digitos:
        return None, None, "PENDENTE_SEM_DOCUMENTO", None

    if len(digitos) < 11:
        ajustado = digitos.zfill(11)
        validacao = validar_documento_brasileiro(ajustado, "CPF")
        if validacao["valido"]:
            return ajustado, "CPF", "CPF_AJUSTADO_ZERO_ESQUERDA", ajustado
        return ajustado, "CPF", "PENDENTE_DOCUMENTO_INVALIDO", ajustado

    if len(digitos) == 11:
        validacao = validar_documento_brasileiro(digitos, "CPF")
        if validacao["valido"]:
            return digitos, "CPF", "CPF_OK", None
        return digitos, "CPF", "PENDENTE_DOCUMENTO_INVALIDO", None

    if len(digitos) == 14:
        validacao = validar_documento_brasileiro(digitos, "CNPJ")
        if validacao["valido"]:
            return digitos, "CNPJ", "CNPJ_OK", None
        return digitos, "CNPJ", "PENDENTE_DOCUMENTO_INVALIDO", None

    return digitos, None, "PENDENTE_DOCUMENTO_COMPRIMENTO_INVALIDO", None


def _mapear_elaborador(raw: str | None) -> tuple[str | None, str | None]:
    mapa = {
        "GABRIEL": "Gabriel Bomfim Bispo",
        "MARINA": "Marina Bastos",
        "EGESILDA": "Egesilda Santos",
        "LEONARDO": "Leonardo Freitas",
        "LEONARDO FREITAS": "Leonardo Freitas",
        "LE": "Adeildes Conceição Cruz",
    }
    chave = _nk(raw)
    return mapa.get(chave), None if chave in mapa else "Elaborador sem mapeamento"


def _mapear_tipo(raw: str | None) -> tuple[str | None, str | None]:
    mapa = {
        "CUSTEIO": "RPV custeio",
        "RPV_CUSTEIO": "RPV custeio",
        "HONORARIOS": "RPV honorários",
        "RPV-HONORARIOS": "RPV honorários",
        "RPV-PESSOAL": "RPV pessoal",
        "PESSOAL": "RPV pessoal",
        "TRABALHISTA": "RPV trabalhista",
        "RPV-TRABALHISTA": "RPV trabalhista",
        "RPV -TRABALHISTA": "RPV trabalhista",
        "PERICIAL": "RPV periciais",
        "PERICIAIS": "RPV periciais",
        "RPV-PERICIAL": "RPV periciais",
        "RPV-PERICIAIS": "RPV periciais",
        "RPV - PERICIAIS": "RPV periciais",
        "RPV-FEDERAL": "RPV federal",
        "GUIA-DE-CUSTAS": "Guia de custas",
        "INDENIZACAO": "Indenização",
        "RPV - INDENIZACAO": "Indenização",
        "RPV-DANOS MORAIS": "Danos Morais",
        "RPV - DANOS MORAIS": "Danos Morais",
        "RPV_DANOS_MORAIS": "Danos Morais",
        "DANOS MORAIS": "Danos Morais",
    }
    chave = _nk(raw)
    destino = mapa.get(chave)
    if not destino:
        return None, "Tipo de RPV sem mapeamento"
    return destino, None


def _mapear_situacao_empenho(raw: str | None) -> tuple[str | None, str | None]:
    mapa = {
        "CONCLUIDA": "Concluída",
        "PAGO": "Pago",
        "PD GERADA - SEFAZ": "PD Gerada - SEFAZ",
        "AGUARDANDO RETORNO BANCO": "Aguardando Retorno Banco",
        "SE AGUARDANDO APROVACAO": "SE Aguardando Aprovação",
        "GUIAS GERADAS": "Guias Geradas",
        "VD  A LIQUIDAR": "VD à Liquidar",
        "VD A LIQUIDAR": "VD à Liquidar",
        "AGUARDANDO ASSINATURA DA OB": "Aguardando Assinatura da OB",
    }
    chave = _nk_compacto(raw)
    if not chave:
        return None, "Situação de empenho ausente"
    if chave == "VD LIQUIDADA":
        return "VD Liquidada", None
    destino = mapa.get(chave)
    if destino:
        return destino, None
    if chave == "PENDENTE":
        return None, "Situação de empenho Pendente exige revisão manual"
    if chave == "EM EXECUCAO":
        return None, "Situação de empenho Em Execução exige revisão manual"
    return None, "Situação de empenho sem mapeamento"


def _inferir_sem_irrf(situacao_imposto_raw: str | None, imposto_raw) -> bool:
    if _nk(situacao_imposto_raw) == "SEM IRRF":
        return True
    imposto = _decimal_planilha(imposto_raw)
    return isinstance(imposto, str) and imposto.casefold() in {"s/if", "sif"}


def _mapear_situacao_imposto(raw: str | None, *, sem_irrf: bool, valor_irrf) -> tuple[str | None, str | None]:
    mapa = {
        "SEM IRRF": "Sem IRRF",
        "AGUARDANDO PGTO OB PRINCIPAL": "Aguardando PGTO OB Principal",
        "CONCLUIDA": "Concluída",
        "PD IRRF - AGUARDANDO SEFAZ": "PD IRRF - Aguardando SEFAZ",
    }
    chave = _nk(raw)
    if chave:
        destino = mapa.get(chave)
        if not destino:
            return None, "Situação de imposto sem mapeamento"
        return destino, None

    if sem_irrf:
        return "Sem IRRF", None

    if valor_irrf is not None:
        return None, "Situação de imposto ausente com IRRF informado"

    return None, "Situação de imposto ausente"


def _mapear_reinf(raw: str | None, *, sem_irrf: bool) -> tuple[str | None, str | None]:
    if sem_irrf:
        return None, None

    chave = _nk(raw)
    if not chave:
        return None, None
    if chave == "PREENCHIDA":
        return "Concluído", None
    if chave == "CONCLUIDO":
        return "Concluído", None
    if chave == "CANCELADO":
        return "Cancelado", None
    return None, "Status REINF sem mapeamento"


def _valor_irrf(raw, *, sem_irrf: bool):
    valor = _decimal_planilha(raw)
    if sem_irrf:
        return None, None if not isinstance(valor, Decimal) else "IRRF numérico com Sem IRRF informado"
    if valor is None:
        return None, None
    if isinstance(valor, Decimal):
        return valor, None
    return None, "Valor de IRRF inválido"


def _valor_bruto(raw):
    valor = _decimal_planilha(raw)
    if isinstance(valor, Decimal):
        return valor, None
    return None, "Valor bruto ausente ou inválido"


def _numero_processo(raw) -> tuple[str | None, str | None]:
    texto = _texto_limpo(raw)
    if not texto:
        return None, "Número do processo ausente"
    return texto, None


def _processo_edoc(raw) -> tuple[str | None, str | None]:
    texto = _texto_limpo(raw)
    if not texto:
        return None, "C.I./processo E-DOC ausente"
    return texto, None


def _nome_beneficiario(raw) -> tuple[str | None, str | None]:
    texto = _texto_limpo(raw)
    if not texto or texto in NOME_PLACEHOLDERS:
        return None, "Nome do beneficiário ausente"
    return texto, None


@dataclass
class LinhaImportacaoRPV:
    linha_original: int
    exercicio: str | None
    elaborador_origem: str | None
    elaborador_destino: str | None
    descricao_origem: str | None
    tipo_rpv_destino: str | None
    processo_edoc: str | None
    nome_beneficiario: str | None
    documento_origem: str | None
    documento_ajustado: str | None
    tipo_documento: str | None
    status_documento: str
    documento_corrigido: str | None
    numero_processo: str | None
    numero_processo_normalizado: str
    data_ci: date | None
    data_pagamento: date | None
    data_pagamento_irrf: date | None
    valor_bruto: Decimal | None
    valor_irrf: Decimal | None
    sem_irrf: bool
    nota_empenho: str | None
    situacao_empenho_origem: str | None
    situacao_empenho_destino: str | None
    situacao_imposto_origem: str | None
    situacao_imposto_destino: str | None
    ordem_bancaria: str | None
    reinf_origem: str | None
    reinf_destino: str | None
    ob_imposto: str | None
    observacoes: str | None
    issues: list[str] = field(default_factory=list)
    duplicado_banco: bool = False
    duplicado_planilha: bool = False
    duplicado_detalhe: str | None = None
    registro_existente_id: int | None = None
    conciliavel_banco: bool = False
    conciliado: bool = False
    conciliacao_detalhe: str | None = None
    importado: bool = False
    registro_id: int | None = None
    erro_importacao: str | None = None

    @property
    def apta_importacao(self) -> bool:
        return not self.issues and not self.duplicado_banco and not self.duplicado_planilha

    def adicionar_issue(self, mensagem: str | None):
        if mensagem and mensagem not in self.issues:
            self.issues.append(mensagem)


def carregar_planilha_rpvs_normais(path: str | Path) -> list[LinhaImportacaoRPV]:
    workbook = load_workbook(Path(path), read_only=True, data_only=True)
    try:
        sheet = workbook.active
        linhas: list[LinhaImportacaoRPV] = []

        for linha_excel, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            descricao = _texto_limpo(row[2])
            if _nk(descricao) in SKIP_DESCRICOES:
                continue

            exercicio = _competencia_planilha(row[0])
            elaborador_destino, erro_elaborador = _mapear_elaborador(row[1])
            tipo_rpv_destino, alerta_tipo = _mapear_tipo(descricao)
            processo_edoc, erro_edoc = _processo_edoc(row[3])
            nome_beneficiario, erro_nome = _nome_beneficiario(row[4])
            documento_ajustado, tipo_documento, status_documento, documento_corrigido = (
                _normalizar_documento_importacao(row[5])
            )
            numero_processo, erro_numero = _numero_processo(row[6])
            numero_processo_normalizado = normalizar_numero_processo(numero_processo or "")
            data_ci = _data_planilha(row[7])
            valor_bruto, erro_valor_bruto = _valor_bruto(row[8])
            sem_irrf = _inferir_sem_irrf(row[12], row[9])
            valor_irrf, erro_valor_irrf = _valor_irrf(row[9], sem_irrf=sem_irrf)
            situacao_empenho_destino, erro_empenho = _mapear_situacao_empenho(row[11])
            situacao_imposto_destino, erro_imposto = _mapear_situacao_imposto(
                row[12],
                sem_irrf=sem_irrf,
                valor_irrf=valor_irrf,
            )
            reinf_destino, erro_reinf = _mapear_reinf(row[14], sem_irrf=sem_irrf)

            competencia_base = _data_base_competencia(exercicio)
            data_pagamento = (
                competencia_base
                if _nk(situacao_empenho_destino) in STATUS_EMPENHO_PAGOS
                else None
            )
            data_pagamento_irrf = (
                competencia_base
                if not sem_irrf and _nk(situacao_imposto_destino) == "CONCLUIDA"
                else None
            )

            linha = LinhaImportacaoRPV(
                linha_original=linha_excel,
                exercicio=exercicio,
                elaborador_origem=_texto_limpo(row[1]),
                elaborador_destino=elaborador_destino,
                descricao_origem=descricao,
                tipo_rpv_destino=tipo_rpv_destino,
                processo_edoc=processo_edoc,
                nome_beneficiario=nome_beneficiario,
                documento_origem=_texto_limpo(row[5]),
                documento_ajustado=documento_ajustado,
                tipo_documento=tipo_documento,
                status_documento=status_documento,
                documento_corrigido=documento_corrigido,
                numero_processo=numero_processo,
                numero_processo_normalizado=numero_processo_normalizado,
                data_ci=data_ci,
                data_pagamento=data_pagamento,
                data_pagamento_irrf=data_pagamento_irrf,
                valor_bruto=valor_bruto,
                valor_irrf=valor_irrf,
                sem_irrf=sem_irrf,
                nota_empenho=_texto_limpo(row[10]),
                situacao_empenho_origem=_texto_limpo(row[11]),
                situacao_empenho_destino=situacao_empenho_destino,
                situacao_imposto_origem=_texto_limpo(row[12]),
                situacao_imposto_destino=situacao_imposto_destino,
                ordem_bancaria=_texto_limpo(row[13]),
                reinf_origem=_texto_limpo(row[14]),
                reinf_destino=reinf_destino,
                ob_imposto=None if sem_irrf else _texto_limpo(row[15]),
                observacoes=_texto_limpo(row[16]),
            )

            if not exercicio:
                linha.adicionar_issue("Exercício ausente ou inválido")
            if not data_ci:
                linha.adicionar_issue("Data da C.I. ausente ou inválida")
            if not numero_processo_normalizado:
                linha.adicionar_issue("Número do processo inválido")
            if status_documento.startswith("PENDENTE"):
                linha.adicionar_issue("Documento pendente")

            for erro in [
                erro_elaborador,
                alerta_tipo,
                erro_edoc,
                erro_nome,
                erro_numero,
                erro_valor_bruto,
                erro_valor_irrf,
                erro_empenho,
                erro_imposto,
                erro_reinf,
            ]:
                linha.adicionar_issue(erro)

            linhas.append(linha)

        return linhas
    finally:
        workbook.close()


def aplicar_bloqueios_duplicidade(linhas: list[LinhaImportacaoRPV]) -> None:
    agrupadas = defaultdict(list)
    for linha in linhas:
        if linha.issues or not linha.numero_processo_normalizado:
            continue
        agrupadas[linha.numero_processo_normalizado].append(linha)

    for grupo in agrupadas.values():
        if len(grupo) <= 1:
            continue
        detalhes = ", ".join(str(item.linha_original) for item in grupo)
        for linha in grupo:
            linha.duplicado_planilha = True
            linha.duplicado_detalhe = f"Processo repetido dentro da planilha nas linhas: {detalhes}"


def aplicar_bloqueios_banco(
    linhas: list[LinhaImportacaoRPV],
    *,
    chaves_existentes: dict[str, list[str]],
) -> None:
    for linha in linhas:
        if linha.issues or linha.duplicado_planilha:
            continue
        detalhes = chaves_existentes.get(linha.numero_processo_normalizado)
        if not detalhes:
            continue
        linha.duplicado_banco = True
        linha.duplicado_detalhe = " | ".join(detalhes[:3])


def _registro_rpv_em_estado_inicial_para_conciliacao(
    registro: RegistroRPV,
    *,
    sem_irrf: bool,
) -> bool:
    status_rpv = _nk(getattr(getattr(registro, "situacao_empenho", None), "nome", None))
    status_irrf = _nk(getattr(getattr(registro, "situacao_imposto", None), "nome", None))

    if status_rpv != "SEM TRATAMENTO":
        return False

    if sem_irrf:
        if status_irrf not in {"SEM TRATAMENTO", "SEM IRRF"}:
            return False
    elif status_irrf != "SEM TRATAMENTO":
        return False

    if any(
        [
            registro.data_pagamento,
            registro.data_pagamento_irrf,
            str(registro.nota_empenho or "").strip(),
            str(registro.ordem_bancaria or "").strip(),
            str(registro.ob_imposto or "").strip(),
            str(registro.reinf_status or "").strip(),
        ]
    ):
        return False

    return True


def marcar_conciliacoes_banco(linhas: list[LinhaImportacaoRPV]) -> None:
    registros_por_processo: dict[str, list[RegistroRPV]] = defaultdict(list)

    registros = RegistroRPV.query.join(Processo).all()
    for registro in registros:
        if not registro.ativo or registro.status_principal_cancelado:
            continue
        chave = normalizar_numero_processo(
            registro.processo.numero_processo if registro.processo else ""
        )
        if chave:
            registros_por_processo[chave].append(registro)

    for linha in linhas:
        if not linha.duplicado_banco or linha.issues or linha.duplicado_planilha:
            continue

        candidatos = registros_por_processo.get(linha.numero_processo_normalizado, [])
        nome_linha = normalizar_nome(linha.nome_beneficiario)
        documento_linha = normalizar_documento(linha.documento_ajustado)

        for registro in candidatos:
            nome_registro = normalizar_nome(registro.nome_beneficiario)
            documento_registro = normalizar_documento(registro.documento_original)
            if nome_registro != nome_linha or documento_registro != documento_linha:
                continue

            linha.registro_existente_id = registro.id
            if _registro_rpv_em_estado_inicial_para_conciliacao(
                registro,
                sem_irrf=linha.sem_irrf,
            ):
                linha.conciliavel_banco = True
                linha.conciliacao_detalhe = (
                    "Registro existente em estado inicial pode ser atualizado pela planilha."
                )
            else:
                linha.conciliacao_detalhe = (
                    "Registro existente ja estava trabalhado; conciliacao automatica nao aplicada."
                )
            break


def coletar_processos_existentes() -> dict[str, list[str]]:
    chaves: dict[str, list[str]] = defaultdict(list)

    registros = RegistroRPV.query.join(Processo).all()
    for registro in registros:
        if not registro.ativo:
            continue
        chave = normalizar_numero_processo(registro.processo.numero_processo if registro.processo else "")
        if not chave:
            continue
        chaves[chave].append(
            "RPV existente "
            f"(id={registro.id}, beneficiario={registro.nome_beneficiario}, "
            f"status={getattr(getattr(registro, 'situacao_empenho', None), 'nome', '-')})"
        )

    itens = DativoItem.query.all()
    for item in itens:
        if not item.ativo:
            continue
        chave = normalizar_numero_processo(item.numero_processo)
        if not chave:
            continue
        chaves[chave].append(
            "Dativo existente "
            f"(id={item.id}, beneficiario={item.nome_beneficiario}, "
            f"grupo={item.grupo})"
        )

    return chaves


def validar_referencias_sistema() -> dict[str, dict[str, object]]:
    usuarios = {_nk(item.nome): item for item in User.query.filter_by(ativo=True).all()}
    tipos = {_nk(item.nome): item for item in TipoRPV.query.filter_by(ativo=True).all()}
    situacoes_empenho = {
        _nk(item.nome): item for item in SituacaoEmpenho.query.filter_by(ativo=True).all()
    }
    situacoes_imposto = {
        _nk(item.nome): item for item in SituacaoImposto.query.filter_by(ativo=True).all()
    }
    return {
        "usuarios": usuarios,
        "tipos": tipos,
        "situacoes_empenho": situacoes_empenho,
        "situacoes_imposto": situacoes_imposto,
    }


def _data_historica_base(linha: LinhaImportacaoRPV) -> datetime:
    base = linha.data_ci or linha.data_pagamento or _data_base_competencia(linha.exercicio)
    base = base or date.today()
    return datetime.combine(base, time(12, 0))


def importar_linhas(linhas: list[LinhaImportacaoRPV]) -> dict[str, int]:
    referencias = validar_referencias_sistema()
    usuarios = referencias["usuarios"]
    tipos = referencias["tipos"]
    situacoes_empenho = referencias["situacoes_empenho"]
    situacoes_imposto = referencias["situacoes_imposto"]

    contadores = Counter()

    for linha in linhas:
        if not linha.apta_importacao:
            continue

        try:
            usuario = usuarios.get(_nk(linha.elaborador_destino))
            tipo = tipos.get(_nk(linha.tipo_rpv_destino))
            situacao_rpv = situacoes_empenho.get(_nk(linha.situacao_empenho_destino))
            situacao_irrf = situacoes_imposto.get(_nk(linha.situacao_imposto_destino))

            if not usuario:
                raise ValueError("Usuário responsável não encontrado no sistema.")
            if not tipo:
                raise ValueError("Tipo de RPV não encontrado no sistema.")
            if not situacao_rpv:
                raise ValueError("Situação de empenho não encontrada no sistema.")
            if not situacao_irrf:
                raise ValueError("Situação de imposto não encontrada no sistema.")

            data_historica = _data_historica_base(linha)

            processo = Processo(
                exercicio=linha.exercicio,
                processo_edoc=linha.processo_edoc,
                numero_processo=linha.numero_processo,
                data_ci=linha.data_ci,
                data_cadastro=data_historica,
                observacoes_gerais=None,
                criado_por_id=usuario.id,
                atualizado_por_id=usuario.id,
                criado_em=data_historica,
                atualizado_em=data_historica,
            )
            db.session.add(processo)
            db.session.flush()

            registro = RegistroRPV(
                processo_id=processo.id,
                elaborador_id=usuario.id,
                tipo_rpv_id=tipo.id,
                nome_beneficiario=linha.nome_beneficiario,
                nome_beneficiario_normalizado="",
                tipo_documento=linha.tipo_documento,
                documento_original=linha.documento_ajustado,
                documento_normalizado="",
                documento_corrigido=linha.documento_corrigido,
                data_pagamento=linha.data_pagamento,
                data_pagamento_irrf=linha.data_pagamento_irrf,
                valor_bruto=linha.valor_bruto,
                valor_irrf=linha.valor_irrf,
                valor_liquido=Decimal("0.00"),
                possui_irrf=False,
                sem_irrf=linha.sem_irrf,
                imposto_texto=None,
                nota_empenho=linha.nota_empenho,
                numero_se=None,
                situacao_empenho_id=situacao_rpv.id,
                situacao_imposto_id=situacao_irrf.id,
                ordem_bancaria=linha.ordem_bancaria,
                reinf_status=linha.reinf_destino,
                ob_imposto=linha.ob_imposto,
                historico_auto="",
                observacoes=linha.observacoes,
                ativo=True,
                criado_por_id=usuario.id,
                atualizado_por_id=usuario.id,
                criado_em=data_historica,
                atualizado_em=data_historica,
            )
            registro.atualizar_campos_derivados()
            registro.gerar_historico_auto(
                processo_edoc=processo.processo_edoc,
                numero_processo=processo.numero_processo,
                descricao=tipo.nome,
                data_ci=processo.data_ci,
            )

            db.session.add(registro)
            db.session.flush()
            historico = registrar_evento(
                entidade_tipo="registro_rpv",
                entidade_id=registro.id,
                usuario_id=usuario.id,
                acao="Carga histórica assistida",
                resumo=f"Importado da planilha histórica linha {linha.linha_original}",
                forcar_registro=True,
            )
            if historico:
                historico.criado_em = data_historica

            db.session.commit()
            linha.importado = True
            linha.registro_id = registro.id
            contadores["importados"] += 1
        except Exception as exc:
            db.session.rollback()
            linha.erro_importacao = str(exc)
            contadores["erros"] += 1

    return dict(contadores)


def reconciliar_registros_existentes(linhas: list[LinhaImportacaoRPV]) -> dict[str, int]:
    referencias = validar_referencias_sistema()
    usuarios = referencias["usuarios"]
    tipos = referencias["tipos"]
    situacoes_empenho = referencias["situacoes_empenho"]
    situacoes_imposto = referencias["situacoes_imposto"]

    contadores = Counter()

    for linha in linhas:
        if not linha.conciliavel_banco or not linha.registro_existente_id:
            continue

        try:
            usuario = usuarios.get(_nk(linha.elaborador_destino))
            tipo = tipos.get(_nk(linha.tipo_rpv_destino))
            situacao_rpv = situacoes_empenho.get(_nk(linha.situacao_empenho_destino))
            situacao_irrf = situacoes_imposto.get(_nk(linha.situacao_imposto_destino))
            registro = db.session.get(RegistroRPV, int(linha.registro_existente_id))

            if not usuario:
                raise ValueError("Usuário responsável não encontrado no sistema.")
            if not tipo:
                raise ValueError("Tipo de RPV não encontrado no sistema.")
            if not situacao_rpv:
                raise ValueError("Situação de empenho não encontrada no sistema.")
            if not situacao_irrf:
                raise ValueError("Situação de imposto não encontrada no sistema.")
            if not registro or not registro.processo:
                raise ValueError("Registro existente não encontrado para conciliação.")

            antes = snapshot_entidade("registro_rpv", registro)

            processo = registro.processo
            processo.exercicio = linha.exercicio
            processo.processo_edoc = linha.processo_edoc
            processo.numero_processo = linha.numero_processo
            processo.data_ci = linha.data_ci
            processo.atualizado_por_id = usuario.id

            registro.elaborador_id = usuario.id
            registro.tipo_rpv_id = tipo.id
            registro.nome_beneficiario = linha.nome_beneficiario
            registro.tipo_documento = linha.tipo_documento
            registro.documento_original = linha.documento_ajustado
            registro.documento_corrigido = linha.documento_corrigido
            registro.data_pagamento = linha.data_pagamento
            registro.data_pagamento_irrf = linha.data_pagamento_irrf
            registro.valor_bruto = linha.valor_bruto
            registro.valor_irrf = linha.valor_irrf
            registro.sem_irrf = linha.sem_irrf
            registro.nota_empenho = linha.nota_empenho
            registro.situacao_empenho_id = situacao_rpv.id
            registro.situacao_imposto_id = situacao_irrf.id
            registro.ordem_bancaria = linha.ordem_bancaria
            registro.reinf_status = linha.reinf_destino
            registro.ob_imposto = linha.ob_imposto
            registro.observacoes = linha.observacoes
            registro.atualizado_por_id = usuario.id
            registro.atualizar_campos_derivados()
            registro.gerar_historico_auto(
                processo_edoc=processo.processo_edoc,
                numero_processo=processo.numero_processo,
                descricao=tipo.nome,
                data_ci=processo.data_ci,
            )

            registrar_evento(
                entidade_tipo="registro_rpv",
                entidade_id=registro.id,
                usuario_id=usuario.id,
                acao="Conciliação de carga histórica",
                antes=antes,
                depois=snapshot_entidade("registro_rpv", registro),
                resumo=f"Registro existente conciliado pela planilha histórica linha {linha.linha_original}",
            )

            db.session.commit()
            linha.conciliado = True
            linha.registro_id = registro.id
            linha.conciliacao_detalhe = "Registro existente atualizado com sucesso pela planilha."
            contadores["conciliados"] += 1
        except Exception as exc:
            db.session.rollback()
            linha.erro_importacao = str(exc)
            contadores["erros"] += 1

    return dict(contadores)


def _linha_para_saida(linha: LinhaImportacaoRPV) -> list[object]:
    return [
        linha.linha_original,
        linha.exercicio,
        linha.elaborador_origem,
        linha.elaborador_destino,
        linha.descricao_origem,
        linha.tipo_rpv_destino,
        linha.processo_edoc,
        linha.nome_beneficiario,
        linha.documento_origem,
        linha.documento_ajustado,
        linha.tipo_documento,
        linha.status_documento,
        linha.documento_corrigido,
        linha.numero_processo,
        linha.data_ci,
        linha.data_pagamento,
        linha.data_pagamento_irrf,
        linha.valor_bruto,
        linha.valor_irrf,
        "Sim" if linha.sem_irrf else "Não",
        linha.nota_empenho,
        linha.situacao_empenho_origem,
        linha.situacao_empenho_destino,
        linha.situacao_imposto_origem,
        linha.situacao_imposto_destino,
        linha.ordem_bancaria,
        linha.reinf_origem,
        linha.reinf_destino,
        linha.ob_imposto,
        linha.observacoes,
        " | ".join(linha.issues),
        "Sim" if linha.duplicado_banco else "Não",
        "Sim" if linha.duplicado_planilha else "Não",
        linha.duplicado_detalhe,
        "Sim" if linha.conciliavel_banco else "Não",
        "Sim" if linha.conciliado else "Não",
        linha.conciliacao_detalhe,
        "Sim" if linha.apta_importacao else "Não",
        "Sim" if linha.importado else "Não",
        linha.registro_id,
        linha.erro_importacao,
    ]


def escrever_relatorios_saida(
    *,
    linhas: list[LinhaImportacaoRPV],
    output_dir: str | Path,
    source_path: str | Path,
    db_path: str | Path,
    import_stats: dict[str, int] | None = None,
) -> Path:
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    headers = [
        "LINHA_ORIGINAL",
        "EXERCICIO",
        "ELABORADOR_ORIGEM",
        "ELABORADOR_DESTINO",
        "DESCRICAO_ORIGEM",
        "TIPO_RPV_DESTINO",
        "PROCESSO_EDOC",
        "NOME_BENEFICIARIO",
        "DOCUMENTO_ORIGEM",
        "DOCUMENTO_AJUSTADO",
        "TIPO_DOCUMENTO",
        "STATUS_DOCUMENTO",
        "DOCUMENTO_CORRIGIDO",
        "NUMERO_PROCESSO",
        "DATA_CI",
        "DATA_PAGAMENTO",
        "DATA_PAGAMENTO_IRRF",
        "VALOR_BRUTO",
        "VALOR_IRRF",
        "SEM_IRRF",
        "NOTA_EMPENHO",
        "SITUACAO_EMPENHO_ORIGEM",
        "SITUACAO_EMPENHO_DESTINO",
        "SITUACAO_IMPOSTO_ORIGEM",
        "SITUACAO_IMPOSTO_DESTINO",
        "ORDEM_BANCARIA",
        "REINF_ORIGEM",
        "REINF_DESTINO",
        "OB_IMPOSTO",
        "OBSERVACOES",
        "PENDENCIAS",
        "DUPLICADO_BANCO",
        "DUPLICADO_PLANILHA",
        "DUPLICADO_DETALHE",
        "CONCILIAVEL_BANCO",
        "CONCILIADO",
        "CONCILIACAO_DETALHE",
        "APTA_IMPORTACAO",
        "IMPORTADO",
        "REGISTRO_ID",
        "ERRO_IMPORTACAO",
    ]

    categorias = {
        "rpvs_normais_nao_dativos_preparado.xlsx": linhas,
        "rpvs_normais_documentos_pendentes.xlsx": [
            linha for linha in linhas if "Documento pendente" in linha.issues
        ],
        "rpvs_normais_pendencias_carga.xlsx": [linha for linha in linhas if linha.issues],
        "rpvs_normais_duplicados_banco.xlsx": [linha for linha in linhas if linha.duplicado_banco],
        "rpvs_normais_conciliados_banco.xlsx": [linha for linha in linhas if linha.conciliado],
        "rpvs_normais_duplicados_planilha.xlsx": [
            linha for linha in linhas if linha.duplicado_planilha
        ],
        "rpvs_normais_aptos_importacao.xlsx": [linha for linha in linhas if linha.apta_importacao],
        "rpvs_normais_importados.xlsx": [linha for linha in linhas if linha.importado],
        "rpvs_normais_erros_importacao.xlsx": [
            linha for linha in linhas if linha.erro_importacao
        ],
    }

    for nome_arquivo, linhas_categoria in categorias.items():
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "dados"
        sheet.append(headers)
        for linha in linhas_categoria:
            sheet.append(_linha_para_saida(linha))
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        workbook.save(output_dir / nome_arquivo)

    contadores = Counter()
    for linha in linhas:
        contadores["total_nao_dativos"] += 1
        if "Documento pendente" in linha.issues:
            contadores["documentos_pendentes"] += 1
        if linha.issues:
            contadores["pendencias_carga"] += 1
        if linha.duplicado_banco:
            contadores["duplicados_banco"] += 1
        if linha.duplicado_planilha:
            contadores["duplicados_planilha"] += 1
        if linha.conciliavel_banco:
            contadores["conciliaveis_banco"] += 1
        if linha.conciliado:
            contadores["conciliados_banco"] += 1
        if linha.apta_importacao:
            contadores["aptos_importacao"] += 1
        if linha.importado:
            contadores["importados"] += 1
        if linha.erro_importacao:
            contadores["erros_importacao"] += 1

    report_path = output_dir / "analise_importacao_rpvs_normais.md"
    linhas_md = [
        "# Analise de importacao de RPVs normais",
        "",
        f"- Origem da planilha: `{Path(source_path)}`",
        f"- Banco alvo: `{Path(db_path)}`",
        f"- Linhas nao-dativos analisadas: {contadores['total_nao_dativos']}",
        f"- Aptas para importacao automatica: {contadores['aptos_importacao']}",
        f"- Documentos pendentes: {contadores['documentos_pendentes']}",
        f"- Pendencias gerais de carga: {contadores['pendencias_carga']}",
        f"- Duplicados contra o banco: {contadores['duplicados_banco']}",
        f"- Duplicados dentro da planilha: {contadores['duplicados_planilha']}",
        f"- Conciliáveis contra o banco: {contadores['conciliaveis_banco']}",
        f"- Conciliados contra o banco: {contadores['conciliados_banco']}",
        f"- Importados nesta execucao: {contadores['importados']}",
        f"- Erros de importacao: {contadores['erros_importacao']}",
        "",
        "## Observacoes",
        "- Registros com processo repetido foram separados para tratamento manual e nao entram automaticamente.",
        "- CPFs com menos de 11 digitos foram ajustados com zero a esquerda.",
        "- CNPJs foram preservados como documento valido.",
        "- O elaborador `Lê` foi mapeado para `Adeildes Conceição Cruz`.",
        "- Status `Concluída` e `Pago` receberam data de pagamento no primeiro dia do mes de referencia.",
    ]
    if import_stats:
        linhas_md.extend(
            [
                "",
                "## Resultado da execucao",
                f"- Importados: {import_stats.get('importados', 0)}",
                f"- Conciliados: {import_stats.get('conciliados', 0)}",
                f"- Erros: {import_stats.get('erros', 0)}",
            ]
        )
    report_path.write_text("\n".join(linhas_md), encoding="utf-8")

    return report_path
